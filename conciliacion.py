from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import xlsxwriter
from fastapi import UploadFile


async def _validar_xlsx(
    archivo: UploadFile,
    tipo: str,
) -> dict:
    nombre = archivo.filename or ""

    if not nombre:
        raise ValueError(f"{tipo}: el archivo no tiene nombre.")

    if not nombre.lower().endswith(".xlsx"):
        raise ValueError(
            f"{tipo}: '{nombre}' debe ser un archivo .xlsx para este POC."
        )

    contenido = await archivo.read()

    if not contenido:
        raise ValueError(f"{tipo}: '{nombre}' está vacío.")

    try:
        libro = openpyxl.load_workbook(
            BytesIO(contenido),
            read_only=True,
            data_only=True,
        )
        hojas = libro.sheetnames

        if not hojas:
            raise ValueError("El libro no contiene hojas.")

        primera_hoja = hojas[0]
        ws = libro[primera_hoja]

        filas_muestra = 0
        columnas_detectadas = 0

        for fila in ws.iter_rows(max_row=20, values_only=True):
            filas_muestra += 1
            columnas_detectadas = max(
                columnas_detectadas,
                len(fila),
            )

        libro.close()

    except Exception as exc:
        raise ValueError(
            f"{tipo}: no se pudo abrir '{nombre}' como Excel: {exc}"
        ) from exc

    return {
        "tipo": tipo,
        "archivo": nombre,
        "primera_hoja": primera_hoja,
        "numero_hojas": len(hojas),
        "columnas_detectadas": columnas_detectadas,
        "filas_muestra": filas_muestra,
        "tamano_bytes": len(contenido),
    }


async def ejecutar_conciliacion(
    bdep: UploadFile,
    sap: UploadFile,
    ep: UploadFile,
    ip: UploadFile,
    re: UploadFile,
    ruta_salida: str | Path,
) -> dict:
    """
    POC end-to-end.

    Recibe los cinco Excel reales desde Copilot, los abre en Python y genera
    resultado_conciliacion.xlsx.

    Aquí se reemplazará después esta validación por el código real de
    conciliación de más de 1000 líneas.
    """

    resultados = [
        await _validar_xlsx(bdep, "BDEP"),
        await _validar_xlsx(sap, "SAP"),
        await _validar_xlsx(ep, "EP"),
        await _validar_xlsx(ip, "IP"),
        await _validar_xlsx(re, "RE"),
    ]

    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    workbook = xlsxwriter.Workbook(ruta_salida)
    try:
        ws_resumen = workbook.add_worksheet("Resumen")
        ws_validacion = workbook.add_worksheet("Validacion")

        header = workbook.add_format({"bold": True})

        # Resumen
        resumen = [
            ("estado", "OK"),
            ("cantidad_archivos", 5),
            (
                "mensaje",
                "Python recibió los cinco Excel en una sola llamada "
                "y generó este archivo.",
            ),
        ]

        ws_resumen.write_row(0, 0, ["Campo", "Valor"], header)
        for idx, (campo, valor) in enumerate(resumen, start=1):
            ws_resumen.write(idx, 0, campo)
            ws_resumen.write(idx, 1, valor)

        ws_resumen.set_column(0, 0, 24)
        ws_resumen.set_column(1, 1, 70)

        # Validación
        columnas = [
            "tipo",
            "archivo",
            "primera_hoja",
            "numero_hojas",
            "columnas_detectadas",
            "filas_muestra",
            "tamano_bytes",
        ]

        ws_validacion.write_row(0, 0, columnas, header)

        for fila_idx, resultado in enumerate(resultados, start=1):
            ws_validacion.write_row(
                fila_idx,
                0,
                [resultado[col] for col in columnas],
            )

        ws_validacion.freeze_panes(1, 0)
        ws_validacion.set_column(0, 0, 12)
        ws_validacion.set_column(1, 2, 28)
        ws_validacion.set_column(3, 6, 20)

    finally:
        workbook.close()

    return {
        "cantidad_archivos": 5,
    }
